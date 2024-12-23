from openpyxl import load_workbook
from testlink import TestlinkAPIClient
import re
from Utils.Inputs.Common_inputs import inputs
from Utils.Inputs.DropDown_options import Options
from Utils.Logger.log import get_logger


class Test:
    tlc = TestlinkAPIClient(inputs.API_URL, inputs.KEY)

    def __init__(self):
        self.logger = get_logger()

    def read_test_case_excel(self, file_path):
        global row
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        headers = {cell.value.strip(): idx for idx, cell in enumerate(sheet[1])}
        print(headers)
        required_columns = [
            "Test Suite", "Nested TestSuite", "Test Case Title", "Requirements", "Summary", "preconditions",
            "Keywords", "status", "importance", "TestCase_execution_type", "exec time",
            "Steps_actions","expected_results", "step_execution_type"
        ]

        for column in required_columns:
            if column not in headers:
                raise ValueError(f"Missing required column: {column}")

        test_case_data = []
        current_test_case = {}
        current_step_number = 1

        for row in sheet.iter_rows(min_row=2, values_only=True):
            category = row[headers["Test Suite"]]
            nested_category = row[headers["Nested TestSuite"]]  # Handle nested test suite
            test_case_title = row[headers["Test Case Title"]]
            requirements = row[headers["Requirements"]]
            summary = row[headers["Summary"]]
            preconditions = row[headers["preconditions"]]
            steps_actions = row[headers["Steps_actions"]]
            keywords = row[headers["Keywords"]]
            status = Options.STATUS_MAPPING.get(row[headers["status"]], 0)
            importance = Options.IMPORTANCE_MAPPING.get(row[headers["importance"]], 0)
            TC_execution_type = Options.EXECUTION_TYPE_MAPPING.get(row[headers["TestCase_execution_type"]])
            exec_time = row[headers["exec time"]]
            expected_results = row[headers["expected_results"]]
            step_execution_type = Options.EXECUTION_TYPE_MAPPING.get(row[headers["step_execution_type"]], 0)

            if test_case_title:
                if current_test_case:
                    test_case_data.append(current_test_case)

                current_test_case = {
                    "Category": category,
                    "Nested Category": nested_category,  # Add nested category
                    "Test Case Title": test_case_title,
                    "Requirements": requirements,
                    "Summary": summary,
                    "preconditions": preconditions,
                    "status": status,
                    "importance": importance,
                    "TC_exec_type": TC_execution_type,
                    "exec_time": exec_time,
                    "Keywords": keywords,
                    "Expected Output": expected_results,
                    "Steps Data": []
                }

                # Reset step number for new test case
                current_step_number = 1

            # Collect the steps for this test case
            if steps_actions:
                steps = [re.sub(r'^\d+\.\s*', '', step.strip()) for step in steps_actions.split("\n") if step.strip()]

                for action in steps:
                    step = {
                        "step_number": current_step_number,
                        "actions": action,
                        "expected_results": expected_results,
                        "execution_type": step_execution_type
                    }
                    current_test_case["Steps Data"].append(step)
                    current_step_number += 1

        if current_test_case:
            test_case_data.append(current_test_case)

        # Print and upload test cases
        for test_case in test_case_data:
            status_value = test_case["status"]
            status_key = [k for k, v in Options.STATUS_MAPPING.items() if v == status_value]

            importance_value = test_case["importance"]
            importance_key = str([k for k, v in Options.IMPORTANCE_MAPPING.items() if v == importance_value])

            exec_type_value = test_case["TC_exec_type"]
            exec_type_key = str([k for k, v in Options.EXECUTION_TYPE_MAPPING.items() if v == exec_type_value])

            self.logger.info(
                f"\nCategory: {test_case['Category']}\n"
                f"Nested Category: {test_case['Nested Category']}\n"
                f"Test Case Title: {test_case['Test Case Title']}\n"
                f"Requirements: {test_case['Requirements']}\n"
                f"Summary: {test_case['Summary']}\n"
                f"Preconditions: {test_case['preconditions']}\n"
                f"Keywords: {test_case['Keywords']}\n"
                f"Status: {status_key}\n"
                f"Importance: {importance_key}\n"
                f"Execution Type: {exec_type_key}\n"
                f"Execution Time: {test_case['exec_time']}\n"
                f"Expected Output: {test_case['Expected Output']}\n"
                f"Steps Data: {test_case['Steps Data']}\n"
                + "-" * 40
            )

            # Upload to TestLink
            self.upload_test_case_to_testlink(test_case)

        return test_case_data

    def upload_test_case_to_testlink(self, test_case):
        parent_suite_id = self.get_or_create_test_suite(test_case['Category'], test_case['Nested Category'])
        project_id = self.get_project_id(inputs.PROJECT_NAME)
        case_name = test_case['Test Case Title']
        summary = f"{test_case['Summary']}"
        preconditions = test_case['preconditions']
        status = test_case['status']
        importance = test_case['importance']
        time = test_case['exec_time']
        TC_execution_type = test_case['TC_exec_type']
        expected_results = test_case['Expected Output']
        steps_list = test_case['Steps Data']
        keywords_list = test_case['Keywords'].split(",") if test_case['Keywords'] else []

        # Check if the test case already exists
        if self.test_case_exists(case_name, parent_suite_id):
            self.logger.info(
                f"Test case '{case_name}' already exists in suite '{test_case['Category']} > {test_case['Nested Category']}'. Skipping creation.")
            return

        try:
            test_case_response = self.tlc.createTestCase(
                testcasename=case_name,
                testsuiteid=parent_suite_id,
                testprojectid=project_id,
                authorlogin="admin",
                summary=summary,
                preconditions=preconditions,
                status=status,
                importance=importance,
                estimatedexecduration=time,
                executiontype=TC_execution_type,
                steps=steps_list,
                expectedresults=expected_results
            )
            print(f"Test case '{case_name}' created successfully.")

            test_case_id = test_case_response[0]['id']
            details_of_testcase = self.tlc.getTestCaseIDByName(case_name)
            test_id = details_of_testcase[0]["id"]
            tc_full_ext_id = self.tlc.getTestCase(testcaseid=test_id)[0]["full_tc_external_id"]

            if keywords_list:
                keywords = [keyword.strip() for keyword in keywords_list if keyword.strip()]
                response_keyw = self.tlc.addTestCaseKeywords({tc_full_ext_id: keywords})
                self.logger.info(f"Keywords added to test case {case_name}:", response_keyw)
                print("keywords added successfully...........\n")

        except Exception as e:
            print(f"Error creating test case '{case_name}': {str(e)}")

    def test_case_exists(self, test_case_name, suite_id):
        # Check if the test case exists in the specified test suite
        test_cases = self.tlc.getTestCasesForTestSuite(suite_id, False, False)
        for test_case in test_cases:
            existing_name = test_case["name"].strip().lower()
            if existing_name == test_case_name.strip().lower():
                return True
        return False

    def get_project_id(self, project_name):
        projects = self.tlc.getProjects()
        for project in projects:
            if project["name"] == project_name:
                return project["id"]
        raise ValueError(f"Project '{project_name}' not found.")

    def get_or_create_test_suite(self, suite_name, nested_suite_name=None):
        project_id = self.get_project_id(inputs.PROJECT_NAME)

        # Split the suite_name to handle parent and nested suite
        suite_parts = suite_name.split(' > ')
        parent_suite_name = suite_parts[0].strip()  # Parent Suite
        nested_suite_name = suite_parts[1].strip() if len(suite_parts) > 1 else nested_suite_name
        print("nneeeeeeeeeeeeeeeee",nested_suite_name)

        # Get the parent suite ID from TestLink
        parent_suite_id = None
        suites = self.tlc.getFirstLevelTestSuitesForTestProject(project_id)
        for suite in suites:
            if isinstance(suite, dict) and suite.get("name") == parent_suite_name:
                parent_suite_id = suite.get("id")
                break

        # If the parent suite doesn't exist, create it
        if not parent_suite_id:
            suite_response = self.tlc.createTestSuite(project_id, parent_suite_name, "Imported test suite")
            print(f"Created Parent Suite: {suite_response}")
            parent_suite_id = suite_response[0]["id"]

        # Now handle the creation or identification of the nested suite
        nested_suite_id = None
        if nested_suite_name:
            # Debug nested suites
            print(f"Fetching nested suites under Parent Suite ID: {parent_suite_id}")
            nested_suites = self.tlc.getTestSuitesForTestSuite(parent_suite_id)
            print("%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%",nested_suites)
            if isinstance(nested_suites, dict):
                if "id" in nested_suites and "name" in nested_suites:
                    if nested_suites["name"] == nested_suite_name:
                        nested_suite_id = nested_suites["id"]
                        print(f"Using existing Nested Suite ID: {nested_suite_id}")
                else:
                    # Handle multiple suites as a dictionary of dictionaries
                    for suite_id, suite_details in nested_suites.items():
                        if suite_details.get("name") == nested_suite_name:
                            nested_suite_id = suite_details.get("id")
                            print(f"Using existing Nested Suite ID: {nested_suite_id}")
                            break

            # If nested suite doesn't exist, create it
            if not nested_suite_id:
                try:
                    nested_suite_response = self.tlc.createTestSuite(
                        project_id, nested_suite_name, "Nested test suite", parentid=parent_suite_id
                    )
                    if 'id' in nested_suite_response[0]:
                        nested_suite_id = nested_suite_response[0]["id"]
                        print(f"Created Nested Suite: {nested_suite_name} with ID: {nested_suite_id}")
                    else:
                        self.logger.error(f"Failed to create nested suite. Response: {nested_suite_response}")
                        return parent_suite_id  # Return parent suite ID if nested suite creation fails
                except Exception as e:
                    self.logger.error(f"Error creating nested suite: {e}")
                    return parent_suite_id  # Return parent suite ID if nested suite creation fails

        # Return the ID of the nested suite (if it exists) or the parent suite ID
        print("Final Nested Suite ID:", nested_suite_id if nested_suite_id else parent_suite_id)
        return nested_suite_id if nested_suite_id else parent_suite_id


if __name__ == '__main__':
    ts = Test()
    ts.read_test_case_excel(inputs.EXCEL_PATH)
